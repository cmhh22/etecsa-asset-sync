"""
Generate demo Excel files for testing the ETECSA Asset Sync workflow.
These files contain fictional data that mimics the real structure.

Run: python generate_demo_data.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================================================================
# 1. AR01 Demo (Finance Department Report)
# ============================================================================
# The real AR01 starts data at row 9 (8 header rows), uses columns 6 and 9

wb_ar01 = Workbook()
ws = wb_ar01.active
ws.title = "AR01"

# Header rows (rows 1-8) — mimicking the real format
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

ws.merge_cells('A1:I1')
ws['A1'] = 'ETECSA - Empresa de Telecomunicaciones de Cuba S.A.'
ws['A1'].font = Font(bold=True, size=14)

ws.merge_cells('A2:I2')
ws['A2'] = 'Cienfuegos Territorial Division'
ws['A2'].font = Font(bold=True, size=12)

ws.merge_cells('A3:I3')
ws['A3'] = 'Report AR01-1 — Tangible Fixed Assets'

ws.merge_cells('A4:I4')
ws['A4'] = 'Finance and Economics Department'

ws.merge_cells('A5:I5')
ws['A5'] = f'Issue date: 2024-01-15'

# Row 6-7 blank
# Row 8 = column headers
headers_ar01 = ['Sequence', 'Account', 'SubAccount', 'Description', 'Area',
                'Inventory_No', 'Value', 'Depreciation', 'Location']

for col_idx, header in enumerate(headers_ar01, 1):
    cell = ws.cell(row=8, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Data rows (starting at row 9) — fictional inventory data
# Municipios de Cienfuegos: Cienfuegos, Rodas, Palmira, Lajas, Cruces,
# Cumanayagua, Abreus, Aguada de Pasajeros

demo_assets = [
    [1, '240', '01', 'PC Desktop HP ProDesk', 'IT', 'INV-CFG-001', 1500.00, 300.00, 'L-001'],
    [2, '240', '01', 'PC Desktop Dell OptiPlex', 'IT', 'INV-CFG-002', 1600.00, 320.00, 'L-001'],
    [3, '240', '01', 'Laptop Lenovo ThinkPad', 'HR', 'INV-CFG-003', 2100.00, 420.00, 'L-002'],
    [4, '240', '01', 'PC Desktop HP ProDesk', 'Finance', 'INV-CFG-004', 1500.00, 300.00, 'L-003'],
    [5, '240', '01', 'Monitor Samsung 24"', 'IT', 'INV-CFG-005', 450.00, 90.00, 'L-001'],
    [6, '240', '01', 'Printer HP LaserJet', 'Management', 'INV-CFG-006', 800.00, 160.00, 'L-004'],
    [7, '240', '01', 'PC Desktop Dell', 'Sales', 'INV-CFG-007', 1500.00, 300.00, 'L-005'],
    [8, '240', '01', 'Laptop HP EliteBook', 'IT', 'INV-CFG-008', 2200.00, 440.00, 'L-006'],
    [9, '240', '01', 'PC Desktop Lenovo', 'HR', 'INV-CFG-009', 1400.00, 280.00, 'L-007'],
    [10, '240', '01', 'Switch Cisco 24p', 'Network', 'INV-CFG-010', 3500.00, 700.00, 'L-008'],
    [11, '240', '01', 'PC Desktop HP', 'Operations', 'INV-CFG-011', 1500.00, 300.00, 'L-009'],
    [12, '240', '01', 'Laptop Dell Latitude', 'Sales', 'INV-CFG-012', 1900.00, 380.00, 'L-010'],
    [13, '240', '01', 'PC Desktop HP ProDesk', 'Planning', 'INV-CFG-013', 1500.00, 300.00, 'L-011'],
    [14, '240', '01', 'UPS APC 1500VA', 'IT', 'INV-CFG-014', 600.00, 120.00, 'L-001'],
    [15, '240', '01', 'PC Desktop Dell', 'Logistics', 'INV-CFG-015', 1500.00, 300.00, 'L-012'],
    [16, '240', '01', 'Laptop Lenovo L15', 'Management', 'INV-CFG-016', 1800.00, 360.00, 'L-004'],
    [17, '240', '01', 'PC Desktop HP', 'Technical', 'INV-CFG-017', 1500.00, 300.00, 'L-013'],
    [18, '240', '01', 'Server Dell R740', 'IT', 'INV-CFG-018', 8500.00, 1700.00, 'L-014'],
    # Assets in AR01 but NOT in database (for anomaly testing)
    [19, '240', '01', 'PC Desktop Lenovo', 'HR', 'INV-CFG-050', 1500.00, 300.00, 'L-002'],
    [20, '240', '01', 'Laptop HP ProBook', 'Sales', 'INV-CFG-051', 1700.00, 340.00, 'L-005'],
]

for row_data in demo_assets:
    ws.append(row_data)

# Adjust column widths (skip merged cells)
from openpyxl.cell.cell import MergedCell
for col in ws.columns:
    non_merged = [c for c in col if not isinstance(c, MergedCell)]
    if non_merged:
        max_length = max((len(str(c.value)) for c in non_merged if c.value), default=10)
        ws.column_dimensions[non_merged[0].column_letter].width = max_length + 3

wb_ar01.save('demo_data/AR01_demo.xlsx')
print("✓ AR01_demo.xlsx generated")


# ============================================================================
# 2. Location Classifier Demo (HR/Locations Classifier)
# ============================================================================

wb_clasif = Workbook()
ws2 = wb_clasif.active
ws2.title = "Classifier"

# Headers (columns 1-7, we use columns 5, 6, 7)
headers_clasif = ['Municipality_Code', 'Municipality', 'Center_Code', 'Cost_Center',
                  'Location_ID', 'Location_Description', 'Building']

for col_idx, header in enumerate(headers_clasif, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Location data — based on real Cienfuegos municipalities
locations = [
    ['01', 'Cienfuegos', 'CC-001', 'DT Cienfuegos', 'L-001', 'Server_Room', 'Central_Bldg_CFG'],
    ['01', 'Cienfuegos', 'CC-001', 'DT Cienfuegos', 'L-002', 'HR_Office', 'Central_Bldg_CFG'],
    ['01', 'Cienfuegos', 'CC-001', 'DT Cienfuegos', 'L-003', 'Finance_Dept', 'Central_Bldg_CFG'],
    ['01', 'Cienfuegos', 'CC-001', 'DT Cienfuegos', 'L-004', 'Management_Office', 'Central_Bldg_CFG'],
    ['01', 'Cienfuegos', 'CC-002', 'Telephone Center CFG', 'L-005', 'Sales_Room', 'TC_Cienfuegos'],
    ['02', 'Rodas', 'CC-003', 'TC Rodas', 'L-006', 'Technical_Office', 'TC_Rodas'],
    ['03', 'Palmira', 'CC-004', 'TC Palmira', 'L-007', 'Admin_Office', 'TC_Palmira'],
    ['04', 'Lajas', 'CC-005', 'TC Lajas', 'L-008', 'Equipment_Room', 'TC_Lajas'],
    ['05', 'Cruces', 'CC-006', 'TC Cruces', 'L-009', 'Operations_Office', 'TC_Cruces'],
    ['06', 'Cumanayagua', 'CC-007', 'TC Cumanayagua', 'L-010', 'Sales_Office', 'TC_Cumanayagua'],
    ['07', 'Abreus', 'CC-008', 'TC Abreus', 'L-011', 'Planning_Office', 'TC_Abreus'],
    ['08', 'Aguada de Pasajeros', 'CC-009', 'TC Aguada', 'L-012', 'Logistics_Office', 'TC_Aguada'],
    ['01', 'Cienfuegos', 'CC-010', 'Provincial Node', 'L-013', 'Technical_Room', 'Provincial_Node'],
    ['01', 'Cienfuegos', 'CC-010', 'Provincial Node', 'L-014', 'DataCenter', 'Provincial_Node'],
]

for row_data in locations:
    ws2.append(row_data)

# Adjust column widths
for col in ws2.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
    ws2.column_dimensions[col[0].column_letter].width = max_length + 3

wb_clasif.save('demo_data/clasificador_demo.xlsx')
print("✓ clasificador_demo.xlsx generated")

print("\n✅ All demo data files generated successfully in demo_data/")
